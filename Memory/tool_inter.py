import openpyxl

output_file = 'D:/work1/c_test_file/test.xlsx'

# 获取输出内容
output_content = '''
Filename                        Regions    Missed Regions     Cover   Functions  Missed Functions  Executed       Lines      Missed Lines     Cover    Branches   Missed Branches     Cover
-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
D:\\work1\\c_test_file\\test.c          17                 2    88.24%           4                 1    75.00%          47                 4    91.49%          10                 1    90.00%
-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
TOTAL                                17                 2    88.24%           4                 1    75.00%          47                 4    91.49%          10                 1    90.00%
'''

# 创建Excel工作簿和工作表
workbook = openpyxl.Workbook()
worksheet = workbook.active

# 将输出内容按行分割，并写入工作表中
lines = output_content.strip().split('\n')
for row_index, line in enumerate(lines, start=1):
    if row_index == 1:  # 处理表格标题
        columns = [value for value in line.split('  ') if value]  # 剔除空值
        for col_index, value in enumerate(columns, start=1):
            worksheet.cell(row=row_index, column=col_index).value = value
    else:
        columns = line.split()
        for col_index, value in enumerate(columns, start=1):
            worksheet.cell(row=row_index, column=col_index).value = value

# 保存Excel文件
workbook.save(output_file)

print(f"Excel文件已生成：{output_file}")
